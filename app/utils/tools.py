#!/usr/bin/env python
# -*- coding: utf-8 -*-
import decimal
import datetime
import types

from flask.json import JSONEncoder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment


class CustomJSONEncoder(JSONEncoder):

    def default(self, obj):
        try:
            if isinstance(obj, datetime.datetime) or isinstance(obj, datetime.date):
                encoded_object = obj.strftime('%Y-%m-%d')
                return encoded_object
            iterable = iter(obj)
        except TypeError:
            pass
        else:
            return list(iterable)
        return JSONEncoder.default(self, obj)


# excel设置范围样式函数
def style_range(ws, cell_range=None, border=None, fill=None, alignment=None, merge_header=False, auto_width=False):
    """
    :param ws: Excel worksheet instance
    :param cell_range:  An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param alignment:  An openpyxl alignment object
    :param merge_header
    :param auto_width
    :return:
    """
    if border is None:
        # 定义边框填充
        bd = Side(border_style='thin', color='000000')
        border = Border(left=bd, right=bd, top=bd, bottom=bd)
    if alignment is None:
        alignment = Alignment(horizontal='center', vertical='center')
    if cell_range is None:
        if merge_header:
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ws.max_column)
            ws.row_dimensions[1].height = 40
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            # 设置每一行的高
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 25
            cell_range = get_column_letter(1) + "2:" + get_column_letter(ws.max_column) + str(
                ws.max_row)
        else:
            cell_range = ws.dimensions

    rows = ws[cell_range]

    column_widths = dict()

    for row in rows:
        for c in row:
            if fill:
                c.fill = fill
            if alignment:
                c.alignment = alignment
            if border:
                c.border = border

    # 列宽自适应
    if auto_width:
        for row in rows:
            for c in row:
                if c.value is not None:
                    if isinstance(c.value, datetime.datetime) or isinstance(c.value, datetime.datetime):
                        cell_width = len(str(c.value)) + 4
                    elif isinstance(c.value, types.UnicodeType):
                        cell_width = len(c.value.encode('utf-8')) + 4
                    else:
                        cell_width = 20
                    if c.column not in column_widths:
                        column_widths.update({c.column: cell_width})
                    else:
                        if cell_width > column_widths.get(c.column, 20):
                            column_widths.update({c.column: cell_width})
        for k, v in column_widths.iteritems():
            ws.column_dimensions[k].width = v


def percent_div(up, down):
    if up == 0 or up is None:
        return 0
    try:
        return round((up / down) * 100, 2)
    except ZeroDivisionError:
        return 0

